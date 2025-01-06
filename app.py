import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

def remove_text(text, phrase):
    return text.replace(phrase, '')


webpage = "https://www.booking.com/searchresults.html?ss=Dehradun&ssne=Dehradun&ssne_untouched=Dehradun&label=vienna-PNVY5BTMRTTJHYXTpLjJCAS553249712605%3Apl%3Ata%3Ap1%3Ap2%3Aac%3Aap%3Aneg%3Afi%3Atikwd-1365191499%3Alp9298496%3Ali%3Adec%3Adm%3Appccp%3DUmFuZG9tSVYkc2RlIyh9YZVcNNsENnH02-pWD53qm9c&aid=306395&lang=en-us&sb=1&src_elem=sb&src=searchresults&dest_id=-2094211&dest_type=city&checkin=2025-01-09&checkout=2025-01-10&group_adults=2&no_rooms=1&group_children=0"


filename = "hotels_in_dehradun.xlsx"


driver = webdriver.Firefox()
driver.get(webpage)


hotels = driver.find_elements(By.CLASS_NAME, "aab71f8e4e")
hotel_prices = driver.find_elements(By.CSS_SELECTOR, "span[data-testid='price-and-discounted-price']")

# Extracting hotel names and prices
hotel_data = []
for hotel, price in zip(hotels, hotel_prices):
    hotel_name = remove_text(hotel.text, "\nOpens in new window")
    hotel_price = price.text
    hotel_data.append((hotel_name.strip(), hotel_price))

driver.quit()

# Creating an Excel file
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Hotel Prices"

ws['A1'] = "Hotel Name"
ws['B1'] = "Price Per Night"

for i, (hotel_name, hotel_price) in enumerate(hotel_data, start=2):
    ws[f'A{i}'] = hotel_name
    ws[f'B{i}'] = hotel_price


wb.save(filename)
print(f"Excel sheet saved as {filename}")
